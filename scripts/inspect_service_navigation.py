#!/usr/bin/env python3
"""Inspect the private service-navigation exports and emit **safe aggregates only**.

This reads the two private Dataverse Excel exports behind the
``service-navigation/`` module — a resource-directory workbook and a
goal-pathway workbook — and reports file hashes, worksheet counts, row counts,
generalized column classifications, presence rates, duplicate counts,
normalized Type counts, aggregate category counts, hierarchy-integrity counts,
and a cross-workbook resource-reference join (aggregate results only).

**No production row value is ever published.** The inspector never prints,
logs, writes, or serializes a cell value — no organization name, program name,
address, phone number, email address, URL, application instruction, Dataverse
GUID, row checksum, modified timestamp, or requirement text. Raw column
headers are withheld (columns are reported only under generalized public field
names), and **source filenames and paths are withheld** from the report (the
exports embed production terminology and export timestamps; files are
identified by role and SHA-256 only). Every output line passes a leakage guard
that rejects email-, GUID-, URL-, phone-, and checksum-shaped text by
construction.

The sources live under ``source-private/service-navigation/`` and are never
committed (see ``AGENTS.md`` / ``SECURITY.md``). Run locally with the sources
present.

Usage:
    python scripts/inspect_service_navigation.py

Exits non-zero with a clear message if a source is missing, unreadable, or
ambiguous.
"""
from __future__ import annotations

import hashlib
import re
import sys
from collections import Counter
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = REPO_ROOT / "source-private" / "service-navigation"

ROLE_DIRECTORY = "resource directory"
ROLE_PATHWAY = "goal pathway"

# --------------------------------------------------------------------------- #
# Output leakage guard. Every printed line must pass. This is the "safe by
# construction" gate: even a future bug that routed a cell value toward output
# would be caught here before anything is emitted.
# --------------------------------------------------------------------------- #

_LEAK_PATTERNS = (
    ("email", re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")),
    ("guid", re.compile(r"\b[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}\b")),
    ("url", re.compile(r"(?:https?|ftp)://|www\.", re.IGNORECASE)),
    ("phone", re.compile(r"\(?\d{3}\)?[ \-.]\d{3}[ \-.]\d{4}|\b\d{10,}\b")),
    # checksum/base64-like run (Dataverse row checksums are long base64 strings)
    ("checksum", re.compile(r"[A-Za-z0-9+/]{24,}={0,2}")),
)


class LeakageError(RuntimeError):
    """Raised when a would-be output line matches a leakage pattern."""


def guard_line(line: str) -> str:
    """Return ``line`` if it is safe to emit; raise ``LeakageError`` otherwise.

    SHA-256 file hashes are the one allowed long-hex token: a line consisting
    of a labelled 64-char lowercase hex digest is permitted (a file hash is an
    aggregate file property, not a row value) — but only when it matches the
    exact ``sha256=<64 hex>`` form this script prints.
    """
    candidate = re.sub(r"\bsha256=[0-9a-f]{64}\b", "sha256=<hash>", line)
    for name, pattern in _LEAK_PATTERNS:
        if pattern.search(candidate):
            raise LeakageError(f"refusing to emit output line matching {name!r} pattern")
    return line


def emit(lines: list[str], text: str) -> None:
    """Append a guarded output line."""
    lines.append(guard_line(text))


# --------------------------------------------------------------------------- #
# Generalized column classification.
#
# Raw source headers are NEVER emitted. Each header is matched (casefolded,
# keyword rules, most-specific first) to a generalized public field name and a
# privacy category. Unrecognized headers surface as ``unclassified-column-<n>``
# so a schema change cannot silently leak a new header name.
# --------------------------------------------------------------------------- #

CAT_INTERNAL = "internal-metadata"      # GUIDs, checksums, timestamps: withheld
CAT_CONTACT = "contact-channel"         # phones, emails: presence only
CAT_LOCATION = "location"               # addresses: presence only
CAT_LINK = "web-link"                   # URLs: presence only
CAT_TAXONOMY = "taxonomy"               # categories/specializations: aggregate only
CAT_TEXT = "descriptive-text"           # free text: presence only
CAT_HIERARCHY = "hierarchy"             # codes, types, parents: aggregate/shape only
CAT_REFERENCE = "record-reference"      # lookups: link counts only
CAT_SIGNAL = "curation-signal"          # ratings/flags: aggregate only

# (keyword, public field name, category) — first match wins, order matters.
_HEADER_RULES: tuple[tuple[str, str, str], ...] = (
    ("checksum", "row_checksum", CAT_INTERNAL),
    ("modified", "modified_timestamp", CAT_INTERNAL),
    ("do not modify", "internal_identifier", CAT_INTERNAL),
    ("phone", "support_contact", CAT_CONTACT),
    ("email", "contact_email", CAT_CONTACT),
    ("address", "location", CAT_LOCATION),
    ("website", "web_link", CAT_LINK),
    ("hour", "hours_availability", CAT_TEXT),
    ("categor", "top_level_categories", CAT_TAXONOMY),
    ("specializ", "specialization", CAT_TAXONOMY),
    ("service", "service_description", CAT_TAXONOMY),
    ("organization", "organization_title", CAT_TEXT),
    ("program", "program_title", CAT_TEXT),
    ("rating", "average_rating", CAT_SIGNAL),
    ("about", "summary", CAT_TEXT),
    ("copy", "copy_target", CAT_REFERENCE),
    ("resource", "linked_resource", CAT_REFERENCE),
    ("type", "step_type", CAT_HIERARCHY),
    ("goal desc", "goal_summary", CAT_TEXT),
    ("why", "rationale", CAT_TEXT),
    ("application", "application_links", CAT_TEXT),
    ("form", "form_links", CAT_TEXT),
    ("phase", "phase", CAT_HIERARCHY),
    ("optional", "optional_step", CAT_SIGNAL),
    ("note", "note", CAT_TEXT),
    ("need", "need_text", CAT_TEXT),
    ("description", "step_summary", CAT_TEXT),
    ("id", "step_code", CAT_HIERARCHY),
    ("name", "step_title", CAT_TEXT),
)

# Marker fields that identify a data sheet's role. A visible sheet must carry
# all markers of exactly one role to be selected.
_ROLE_MARKERS = {
    ROLE_DIRECTORY: frozenset({"organization_title", "top_level_categories", "service_description"}),
    ROLE_PATHWAY: frozenset({"step_type", "step_code", "linked_resource"}),
}


def classify_header(header: object, position: int) -> tuple[str, str]:
    """Map a raw header to ``(public_field_name, category)``.

    The raw header text never leaves this function. Unmatched or empty headers
    classify as ``unclassified-column-<position>`` / descriptive-text so new
    source columns are visible without being named.
    """
    text = str(header).casefold().strip() if header is not None else ""
    if text:
        for keyword, public_name, category in _HEADER_RULES:
            if keyword in text:
                return public_name, category
    return f"unclassified-column-{position}", CAT_TEXT


# --------------------------------------------------------------------------- #
# Normalization + shape helpers (deterministic; documented in metrics.md).
# --------------------------------------------------------------------------- #

def normalize_value(value: object) -> str | None:
    """Trim, collapse internal whitespace, casefold. ``None`` when blank."""
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).strip())
    return text.casefold() or None


def is_blank_row(row: tuple) -> bool:
    """True when every cell is empty or whitespace-only."""
    return all(cell is None or str(cell).strip() == "" for cell in row)


def code_shape(value: object, max_len: int = 16) -> str:
    """Reduce a hierarchy code to a pattern shape: digits -> ``#``, letters -> ``A``.

    Whitespace runs collapse to one space and the shape is truncated to
    ``max_len`` so free text accidentally present in a code column cannot leak
    structure. Blank cells report ``<empty>``.
    """
    if value is None or str(value).strip() == "":
        return "<empty>"
    shape = re.sub(r"\s+", " ", str(value).strip())
    shape = re.sub(r"[0-9]+", "#", shape)
    shape = re.sub(r"[A-Za-z]+", "A", shape)
    return shape[:max_len] + ("~" if len(shape) > max_len else "")


# A structurally parseable hierarchy code: dotted numeric segments with an
# optional single trailing lowercase letter (an observed source shape whose
# semantics are NOT interpreted here).
_CODE_PATTERN = re.compile(r"^(\d+(?:\.\d+)*)([a-z])?$")


def parse_code(value: object) -> tuple[str, ...] | None:
    """Return the numeric segments of a hierarchy code, or ``None``.

    ``None`` means blank or structurally unparseable. A trailing lowercase
    letter is accepted and dropped for parentage math (shape is still reported
    separately); its meaning is not assumed.
    """
    if value is None:
        return None
    text = str(value).strip().casefold()
    if not text:
        return None
    match = _CODE_PATTERN.match(text)
    if not match:
        return None
    return tuple(match.group(1).split("."))


def sha256_of(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


# --------------------------------------------------------------------------- #
# Workbook loading + worksheet selection. Functions accept explicit paths so
# tests can run them against entirely invented fixture workbooks.
# --------------------------------------------------------------------------- #

def _sheet_role(header: tuple) -> str | None:
    """Role a header row matches, or ``None``. Matching both roles is treated
    as no match (a sheet cannot be both exports)."""
    names = {classify_header(raw, position)[0] for position, raw in enumerate(header)}
    matches = [role for role, markers in _ROLE_MARKERS.items() if markers <= names]
    return matches[0] if len(matches) == 1 else None


def load_export(path: Path) -> dict:
    """Load a private export workbook and select its data worksheet.

    Selection method (reported in the output): the data sheet must be
    **visible** (hidden re-import/metadata sheets are excluded by state, not
    by size) and its header must classify to the expected generalized schema
    of exactly one role (directory or pathway). If no visible sheet matches,
    or more than one matches, a ``ValueError`` explains the ambiguity —
    without naming the file or its sheets.

    Returns ``{role, sheet_count, header, rows, blank_excluded}``.
    """
    from openpyxl import load_workbook  # deferred: tests import this module cheaply

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        candidates: list[tuple[str, str, list[tuple]]] = []  # (sheet, role, rows)
        for name in workbook.sheetnames:
            sheet = workbook[name]
            if getattr(sheet, "sheet_state", "visible") != "visible":
                continue
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            role = _sheet_role(rows[0])
            if role is not None:
                candidates.append((name, role, rows))
        sheet_count = len(workbook.sheetnames)
    finally:
        workbook.close()

    if not candidates:
        raise ValueError(
            "no visible worksheet matches the expected directory or pathway "
            "generalized schema (hidden/metadata sheets are excluded)"
        )
    if len(candidates) > 1:
        raise ValueError(
            f"ambiguous workbook: {len(candidates)} visible worksheets match an "
            "expected generalized schema; expected exactly one data sheet"
        )
    _, role, rows = candidates[0]
    header, data = rows[0], rows[1:]
    kept = [row for row in data if not is_blank_row(row)]
    return {
        "role": role,
        "sheet_count": sheet_count,
        "header": header,
        "rows": kept,
        "blank_excluded": len(data) - len(kept),
    }


def column_index(header: tuple, public_name: str) -> int | None:
    """Index of the column whose generalized classification is ``public_name``."""
    for position, raw in enumerate(header):
        name, _ = classify_header(raw, position)
        if name == public_name:
            return position
    return None


def _cell(row: tuple, col: int | None) -> object:
    return row[col] if col is not None and col < len(row) else None


def presence_counts(header: tuple, rows: list[tuple]) -> list[tuple[str, str, int]]:
    """Per-column ``(public_field, category, populated_count)`` in header order."""
    out = []
    for position, raw in enumerate(header):
        name, category = classify_header(raw, position)
        populated = sum(
            1 for row in rows
            if position < len(row) and row[position] is not None and str(row[position]).strip()
        )
        out.append((name, category, populated))
    return out


# --------------------------------------------------------------------------- #
# Directory analysis.
# --------------------------------------------------------------------------- #

def analyze_directory(header: tuple, rows: list[tuple]) -> dict:
    """Safe aggregates for the resource-directory sheet. No values returned."""
    org_col = column_index(header, "organization_title")
    svc_col = column_index(header, "service_description")
    cat_col = column_index(header, "top_level_categories")
    prog_col = column_index(header, "program_title")

    orgs = [normalize_value(_cell(r, org_col)) for r in rows]
    org_values = [o for o in orgs if o]
    org_counter = Counter(org_values)
    repeated = {k for k, v in org_counter.items() if v > 1}

    # Of the repeated organization values, how many carry more than one
    # distinct normalized program title? (Counted, never named.)
    programs_by_org: dict[str, set] = {}
    for r in rows:
        org = normalize_value(_cell(r, org_col))
        if org in repeated:
            programs_by_org.setdefault(org, set()).add(normalize_value(_cell(r, prog_col)))
    repeated_with_multiple_programs = sum(
        1 for titles in programs_by_org.values() if len(titles - {None}) > 1
    )

    services_trim = {str(_cell(r, svc_col)).strip() for r in rows
                     if _cell(r, svc_col) is not None and str(_cell(r, svc_col)).strip()}
    services_norm = {normalize_value(_cell(r, svc_col)) for r in rows
                     if normalize_value(_cell(r, svc_col))}

    category_tokens: set[str] = set()
    multi_category_rows = 0
    categorized_rows = 0
    for r in rows:
        raw = _cell(r, cat_col)
        if raw is None or not str(raw).strip():
            continue
        categorized_rows += 1
        tokens = [t for t in (normalize_value(part) for part in str(raw).split(";")) if t]
        if len(tokens) > 1:
            multi_category_rows += 1
        category_tokens.update(tokens)

    program_values = [normalize_value(_cell(r, prog_col)) for r in rows]
    program_populated = sum(1 for p in program_values if p)
    programs = {p for p in program_values if p}

    return {
        "rows": len(rows),
        "org_populated": len(org_values),
        "org_distinct": len(org_counter),
        "org_repeated": len(repeated),
        "org_repeated_extra_rows": sum(org_counter[k] - 1 for k in repeated),
        "org_repeated_with_multiple_programs": repeated_with_multiple_programs,
        "service_populated": sum(1 for r in rows if _cell(r, svc_col) is not None and str(_cell(r, svc_col)).strip()),
        "service_distinct_trim": len(services_trim),
        "service_distinct_normalized": len(services_norm),
        "category_populated": categorized_rows,
        "category_distinct": len(category_tokens),
        "category_multi_rows": multi_category_rows,
        "program_populated": program_populated,
        "program_distinct": len(programs),
        "presence": presence_counts(header, rows),
    }


# --------------------------------------------------------------------------- #
# Pathway analysis.
# --------------------------------------------------------------------------- #

# Public labels for the three hierarchy levels. The source's middle label is
# plural; normalization maps either form to the singular public label.
_TYPE_LABELS = {
    "goal": "Goal",
    "action item": "Action Item",
    "action items": "Action Item",
    "need": "Need",
}
_EXPECTED_DEPTH = {"Goal": 1, "Action Item": 2, "Need": 3}


def analyze_hierarchy(typed_codes: list[tuple[str, object]]) -> dict:
    """Hierarchy-integrity aggregates from ``(type_label, raw_code)`` pairs.

    Counts only — actual codes are never returned. A code is *valid* for its
    level when it parses and its numeric depth matches the level. Parentage is
    checked by numeric-prefix among valid codes (the production parent
    mechanism is a lookup field not present in the export; this checks the
    code convention only).
    """
    valid_codes: dict[str, set] = {"Goal": set(), "Action Item": set(), "Need": set()}
    coded = Counter()
    blank = Counter()
    malformed = Counter()
    all_valid_normalized: list[str] = []

    entries: list[tuple[str, tuple[str, ...]]] = []
    for label, raw in typed_codes:
        if label not in _EXPECTED_DEPTH:
            continue
        if raw is None or not str(raw).strip():
            blank[label] += 1
            continue
        segments = parse_code(raw)
        if segments is None or len(segments) != _EXPECTED_DEPTH[label]:
            malformed[label] += 1
            continue
        coded[label] += 1
        valid_codes[label].add(segments)
        all_valid_normalized.append(str(raw).strip().casefold())
        entries.append((label, segments))

    duplicate_codes = sum(1 for _, count in Counter(all_valid_normalized).items() if count > 1)

    action_parent_ok = sum(1 for label, seg in entries
                           if label == "Action Item" and seg[:1] in valid_codes["Goal"])
    action_parent_missing = coded["Action Item"] - action_parent_ok
    need_parent_ok = sum(1 for label, seg in entries
                         if label == "Need" and seg[:2] in valid_codes["Action Item"])
    need_parent_missing = coded["Need"] - need_parent_ok

    return {
        "coded": dict(coded),
        "blank": dict(blank),
        "malformed": dict(malformed),
        "duplicate_codes": duplicate_codes,
        "action_parent_ok": action_parent_ok,
        "action_parent_missing": action_parent_missing,
        "need_parent_ok": need_parent_ok,
        "need_parent_missing": need_parent_missing,
    }


def analyze_pathway(header: tuple, rows: list[tuple]) -> dict:
    """Safe aggregates for the goal-pathway sheet. No values returned."""
    type_col = column_index(header, "step_type")
    code_col = column_index(header, "step_code")
    link_col = column_index(header, "linked_resource")
    opt_col = column_index(header, "optional_step")
    phase_col = column_index(header, "phase")

    type_counts: Counter[str] = Counter()
    shapes_by_type: dict[str, Counter] = {}
    linked_by_type: Counter[str] = Counter()
    linked_targets: set[str] = set()
    top_prefixes: set[str] = set()
    unrecognized_types = 0
    typed_codes: list[tuple[str, object]] = []

    for r in rows:
        label = _TYPE_LABELS.get(normalize_value(_cell(r, type_col)) or "")
        if label is None:
            unrecognized_types += 1
            label = "<unrecognized>"
        type_counts[label] += 1
        raw_code = _cell(r, code_col)
        shapes_by_type.setdefault(label, Counter())[code_shape(raw_code)] += 1
        typed_codes.append((label, raw_code))
        link = normalize_value(_cell(r, link_col))
        if link:
            linked_by_type[label] += 1
            linked_targets.add(link)
        segments = parse_code(raw_code)
        if segments:
            top_prefixes.add(segments[0])

    # Reconciliation is real, not tautological: it holds only when every row
    # carries a recognized type AND the three level counts sum to the total.
    recognized_total = sum(v for k, v in type_counts.items() if k != "<unrecognized>")
    reconciles = unrecognized_types == 0 and recognized_total == len(rows)

    # Optional-step states are reported separately; blank is a third state
    # ("unspecified"), NOT folded into "no" or "required".
    optional = Counter(normalize_value(_cell(r, opt_col)) for r in rows)
    optional_yes = optional.get("yes", 0)
    optional_no = optional.get("no", 0)
    optional_unspecified = len(rows) - optional_yes - optional_no

    return {
        "rows": len(rows),
        "type_counts": dict(type_counts),
        "unrecognized_types": unrecognized_types,
        "reconciles": reconciles,
        "code_shapes_by_type": {k: dict(v) for k, v in shapes_by_type.items()},
        "top_level_codes": len(top_prefixes),
        "hierarchy": analyze_hierarchy(typed_codes),
        "linked_rows": sum(linked_by_type.values()),
        "linked_by_type": dict(linked_by_type),
        "linked_distinct_resources": len(linked_targets),
        "linked_values": linked_targets,  # normalized values for the private join ONLY — never rendered
        "optional_yes": optional_yes,
        "optional_no": optional_no,
        "optional_unspecified": optional_unspecified,
        "phase_populated": sum(1 for r in rows if _cell(r, phase_col) is not None and str(_cell(r, phase_col)).strip()),
        "presence": presence_counts(header, rows),
    }


# --------------------------------------------------------------------------- #
# Cross-workbook join: pathway resource references vs. directory rows.
# --------------------------------------------------------------------------- #

def join_resource_references(directory_header: tuple, directory_rows: list[tuple],
                             linked_values: set[str]) -> dict:
    """Deterministic normalized join of pathway resource references against the
    directory. The pathway lookup displays the referenced record's primary
    name, which corresponds to the directory's program-title column, so the
    join key is the normalized program title.

    Emits **counts only**: matched (reference hits exactly one directory row),
    ambiguous (more than one), unmatched (none), and the denominator
    definition. The values being matched are never returned or rendered.
    """
    prog_col = column_index(directory_header, "program_title")
    title_counts = Counter(
        normalize_value(_cell(r, prog_col)) for r in directory_rows
        if normalize_value(_cell(r, prog_col))
    )
    matched = sum(1 for ref in linked_values if title_counts.get(ref, 0) == 1)
    ambiguous = sum(1 for ref in linked_values if title_counts.get(ref, 0) > 1)
    unmatched = len(linked_values) - matched - ambiguous
    return {
        "distinct_references": len(linked_values),
        "matched": matched,
        "ambiguous": ambiguous,
        "unmatched": unmatched,
        "denominator_rows": len(directory_rows),
        "denominator_definition": "analyzed directory rows; join key = normalized program title (the lookup's primary name)",
    }


# --------------------------------------------------------------------------- #
# Report rendering — the only path to stdout; every line is guarded.
# Source filenames, paths, and embedded export timestamps are never rendered:
# files are identified by ROLE and SHA-256 only.
# --------------------------------------------------------------------------- #

def render_file_report(path: Path, export: dict, analysis: dict) -> list[str]:
    lines: list[str] = []
    emit(lines, f"== Source role: {export['role']} ==")
    emit(lines, f"  sha256={sha256_of(path)}")
    emit(lines, f"  worksheets: {export['sheet_count']} total; selected the single visible "
                "sheet matching the expected generalized schema")
    emit(lines, "  (hidden re-import/metadata sheets excluded by visibility and schema)")
    emit(lines, f"  analyzed rows: {analysis['rows']} (blank rows excluded: {export['blank_excluded']})")
    emit(lines, "  column structure (generalized names only; raw headers withheld):")
    for name, category, populated in analysis["presence"]:
        emit(lines, f"    {name:24} {category:20} populated {populated}/{analysis['rows']}")
    return lines


def render_directory_aggregates(a: dict) -> list[str]:
    lines: list[str] = []
    emit(lines, "  aggregates:")
    emit(lines, f"    distinct organizations (normalized): {a['org_distinct']} "
                f"of {a['org_populated']} populated")
    emit(lines, f"    organization values appearing on multiple rows: {a['org_repeated']} "
                f"(+{a['org_repeated_extra_rows']} extra rows)")
    emit(lines, f"    repeated organizations with >1 distinct program title: "
                f"{a['org_repeated_with_multiple_programs']}")
    emit(lines, f"    program-title cells populated: {a['program_populated']}; "
                f"distinct normalized program titles: {a['program_distinct']}")
    emit(lines, f"    distinct service descriptions: {a['service_distinct_trim']} trim-only / "
                f"{a['service_distinct_normalized']} normalized, of {a['service_populated']} populated")
    emit(lines, f"    distinct top-level categories: {a['category_distinct']} "
                f"(multi-category rows: {a['category_multi_rows']}, categorized rows: {a['category_populated']})")
    return lines


def render_pathway_aggregates(a: dict) -> list[str]:
    lines: list[str] = []
    hierarchy = a["hierarchy"]
    emit(lines, "  aggregates:")
    parts = ", ".join(f"{k}: {v}" for k, v in sorted(a["type_counts"].items()))
    emit(lines, f"    step types (normalized): {parts}")
    if a["reconciles"]:
        emit(lines, f"    reconciliation: PASS (all rows recognized; level counts sum to {a['rows']})")
    else:
        emit(lines, f"    reconciliation: FAIL ({a['unrecognized_types']} unrecognized row(s); "
                    f"recognized levels must sum to {a['rows']})")
    for level, shapes in sorted(a["code_shapes_by_type"].items()):
        shaped = ", ".join(f"{shape} x{count}" for shape, count in sorted(shapes.items()))
        emit(lines, f"    hierarchy-code shapes [{level}]: {shaped}")
    emit(lines, "  hierarchy integrity (code convention; counts only):")
    for level in ("Goal", "Action Item", "Need"):
        emit(lines, f"    {level}: coded {hierarchy['coded'].get(level, 0)}, "
                    f"blank {hierarchy['blank'].get(level, 0)}, "
                    f"malformed {hierarchy['malformed'].get(level, 0)}")
    emit(lines, f"    duplicate valid codes: {hierarchy['duplicate_codes']}")
    emit(lines, f"    Action Items with existing Goal parent code: {hierarchy['action_parent_ok']} "
                f"(missing: {hierarchy['action_parent_missing']})")
    emit(lines, f"    Needs with existing Action Item parent code: {hierarchy['need_parent_ok']} "
                f"(missing: {hierarchy['need_parent_missing']})")
    emit(lines, f"    distinct top-level pathway codes: {a['top_level_codes']}")
    emit(lines, f"    rows with a resource reference: {a['linked_rows']} "
                f"({', '.join(f'{k} {v}' for k, v in sorted(a['linked_by_type'].items()))})")
    emit(lines, f"    distinct normalized reference values: {a['linked_distinct_resources']}")
    emit(lines, f"    optional-step states: yes {a['optional_yes']} / no {a['optional_no']} / "
                f"blank-unspecified {a['optional_unspecified']}")
    emit(lines, f"    phase populated: {a['phase_populated']}")
    return lines


def render_join(join: dict) -> list[str]:
    lines: list[str] = []
    emit(lines, "== Cross-workbook resource-reference join ==")
    emit(lines, f"  denominator: {join['denominator_rows']} "
                f"({join['denominator_definition']})")
    emit(lines, f"  distinct pathway references: {join['distinct_references']}")
    emit(lines, f"  matched exactly one directory row: {join['matched']}")
    emit(lines, f"  ambiguous (matched more than one row): {join['ambiguous']}")
    emit(lines, f"  unmatched: {join['unmatched']}")
    return lines


def find_sources(source_dir: Path) -> dict[str, Path]:
    """Locate the two workbook exports and identify each by role.

    Raises ``FileNotFoundError`` unless exactly one directory-role and one
    pathway-role workbook are present. Error text never includes filenames.
    """
    books = sorted(source_dir.glob("*.xlsx"))
    if len(books) != 2:
        raise FileNotFoundError(
            f"expected exactly 2 .xlsx exports in {source_dir} "
            f"(found {len(books)}). Place the private service-navigation "
            "exports there and re-run."
        )
    roles: dict[str, Path] = {}
    for book in books:
        role = load_export(book)["role"]
        if role in roles:
            raise FileNotFoundError(f"both workbooks match the same role ({role})")
        roles[role] = book
    if set(roles) != {ROLE_DIRECTORY, ROLE_PATHWAY}:
        raise FileNotFoundError("could not identify one directory and one pathway workbook")
    return roles


def main() -> int:
    try:
        roles = find_sources(SOURCE_DIR)
        directory_path = roles[ROLE_DIRECTORY]
        pathway_path = roles[ROLE_PATHWAY]
        directory_export = load_export(directory_path)
        pathway_export = load_export(pathway_path)
    except (FileNotFoundError, ValueError) as exc:
        sys.stderr.write(f"ERROR: {exc}\n")
        return 1

    directory = analyze_directory(directory_export["header"], directory_export["rows"])
    pathway = analyze_pathway(pathway_export["header"], pathway_export["rows"])
    join = join_resource_references(directory_export["header"], directory_export["rows"],
                                    pathway["linked_values"])

    lines: list[str] = []
    emit(lines, "No production row value is published. Sources are identified by role and")
    emit(lines, "hash only; filenames, paths, headers, and cell values are withheld.")
    emit(lines, "")
    lines += render_file_report(directory_path, directory_export, directory)
    lines += render_directory_aggregates(directory)
    emit(lines, "")
    lines += render_file_report(pathway_path, pathway_export, pathway)
    lines += render_pathway_aggregates(pathway)
    emit(lines, "")
    lines += render_join(join)

    print("\n".join(lines))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
